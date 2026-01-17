import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, status
from typing import List, Optional
from app.services.device_inventory_service import DeviceInventoryService
from app.schema.device_inventory_schema import (DeviceInventoryCreate, DeviceInventoryUpdate,
                                                DeviceInventoryInDB,FilterSchema,VendorSchema,DeviceTypeSchema,site_filter)
from app.core.dependencies import get_db
from fastapi.responses import FileResponse
from app.schema.device_inventory_schema import Custom_Response_Inventory,modelCreate
from app.core.dependencies import get_current_active_user
from app.model.user import User
from dependency_injector.wiring import Provide, inject
from app.core.container import Container
from app.schema.site_schema import CustomResponse
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/device_inventory", tags=["Device Inventory"])



@router.post("/devicetype_count", response_model=CustomResponse)
@inject
def device_type(
        model_data:modelCreate,
        current_user: User = Depends(get_current_active_user),
        device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    devices = device_inventory_service.get_device_types(model_data)

    return CustomResponse(
        message="Fetched devices type count successfully",
        data=devices,
        status_code=200
    )

@router.post("/get_device_nature", response_model=CustomResponse)
@inject
def get_count(
        device_nature: modelCreate,
        current_user: User = Depends(get_current_active_user),
        device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    models = device_inventory_service.get_device_nature_data(device_nature)

    return CustomResponse(
        message="Fetched model data successfully",
        data=models,
        status_code=200
    )

@router.get("/get_count", response_model=CustomResponse)
@inject
def get_count(
        current_user: User = Depends(get_current_active_user),
        device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    models = device_inventory_service.get_inventory_counts_data()

    return CustomResponse(
        message="Fetched model data successfully",
        data=models,
        status_code=200
    )


@router.get("/get_vendors", response_model=CustomResponse)
@inject
def get_vendors(
        site_id: Optional[int] = None,
        rack_id: Optional[int]=None,
        current_user: User = Depends(get_current_active_user),
        device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    vendors = device_inventory_service.get_all_vendors(site_id,rack_id)

    return CustomResponse(
        message="Fetched vendors data successfully",
        data=vendors,
        status_code=200
    )


@router.get("/get_vendor_count", response_model=CustomResponse)
@inject
def get_vendor_count(

        current_user: User = Depends(get_current_active_user),
        device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    models = device_inventory_service.get_vendor_counts_data()

    return CustomResponse(
        message="Fetched vendor data successfully",
        data=models,
        status_code=200
    )

@router.post("/get_model_Count", response_model=CustomResponse)
@inject
def get_model_names(
        model_data:modelCreate,
        current_user: User = Depends(get_current_active_user),
        device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    models = device_inventory_service.get_devices_models(model_data)

    return CustomResponse(
        message="Fetched model data successfully",
        data=models,
        status_code=200
    )


@router.post("/get_all_device_inventory", response_model=Custom_Response_Inventory[List[DeviceInventoryInDB]])
@inject
def get_all_devices(page:int=None,
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    devices = device_inventory_service.get_all_devices(page)

    return Custom_Response_Inventory(
        message="Fetched all devices successfully",
        data=devices,
        status_code=200
    )
@router.post("/get_all_device_inventory_with_filter", response_model=Custom_Response_Inventory[List[DeviceInventoryInDB]])
@inject
def get_all_devices(filter_data:FilterSchema,

    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    devices = device_inventory_service.get_all_devices_test(filter_data)

    return Custom_Response_Inventory(
        message="Fetched all devices successfully",
        data=devices,
        status_code=200
    )

@router.get("get_device_inventory_by_id/{device_id}", response_model=Custom_Response_Inventory[DeviceInventoryInDB])
@inject
def get_device_by_id(
    device_id: int, 
    current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    device = device_inventory_service.get_device_by_id(device_id)
    
    return Custom_Response_Inventory(
        message="Fetched device successfully",
        data=device,
        status_code=200
    )


@router.post("/create_device_inventory", response_model=Custom_Response_Inventory[DeviceInventoryInDB])
@inject
def create_device(
    device_data: DeviceInventoryCreate,
    current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    new_device = device_inventory_service.create_device(device_data)
    
    return Custom_Response_Inventory(
        message="Device created successfully",
        data=new_device,
        status_code=201
    )

@router.put("update_device_inventory/{device_id}", response_model=Custom_Response_Inventory[DeviceInventoryInDB])
@inject
def update_device(
    device_id: int, 
    device_data: DeviceInventoryUpdate,
    current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    updated_device = device_inventory_service.update_device(device_id, device_data)
    
    return Custom_Response_Inventory(
        message="Device updated successfully",
        data=updated_device,
        status_code=200
    )


@router.delete("delete_device_inventory/{device_id}", response_model=Custom_Response_Inventory[None])
@inject
def delete_device(
    device_id: int,
    current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    device_inventory_service.delete_device(device_id)
    
    return Custom_Response_Inventory(
        message="Device deleted successfully",
        data=None,
        status_code=200
    )


@router.get("/get_device_inventory_with_power_utilization", response_model=Custom_Response_Inventory[List[DeviceInventoryInDB]])
@inject
def get_device_inventory_with_power_utilization(
    current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    devices = device_inventory_service.get_device_inventory_with_power_utilization()
    
    return Custom_Response_Inventory(
        message="Fetched all devices with power utilization",
        data=devices,
        status_code=200
    )



# Created by Ahmed
@router.get("/chasis", response_model=CustomResponse)
@inject
def chasis(
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    chasis = device_inventory_service.chasis()
    
    return CustomResponse(
        message="Fetched all chassis successfully",
        data=chasis,
        status_code=status.HTTP_200_OK
    )
    
    
    
@router.get("/chasis", response_model=CustomResponse)
@inject
def chasis(
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    chasis = device_inventory_service.chasis()
    
    return CustomResponse(
        message="Fetched all chassis successfully",
        data=chasis,
        status_code=status.HTTP_200_OK
    )
    
    

@router.get("/modules", response_model=CustomResponse)
@inject
def modules(
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    modules = device_inventory_service.modules()
    
    return CustomResponse(
        message="Fetched all modules successfully",
        data=modules,
        status_code=status.HTTP_200_OK
    )
    
    
    
@router.get("/powerSupply", response_model=CustomResponse)
@inject
def power_supply(
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    power_supply = device_inventory_service.power_supply()
    
    return CustomResponse(
        message="Fetched all Power Supply successfully",
        data=power_supply,
        status_code=status.HTTP_200_OK
    )
    
    
    
@router.get("/fans", response_model=CustomResponse)
@inject
def fans(
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    fans = device_inventory_service.fans()
    
    return CustomResponse(
        message="Fetched all Fans successfully",
        data=fans,
        status_code=status.HTTP_200_OK
    )


@router.post("/deviceLastPowerUtiization", response_model=List)
@inject
def device_power(
    apic_api: str,
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    return device_inventory_service.device_power(apic_api)



@router.post("/get_specific_device_inventory", response_model=CustomResponse)
@inject
def get_spcific_devices(
    device_ip: str,
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    devices = device_inventory_service.get_spcific_devices(device_ip)

    return CustomResponse(
        message="Fetched device data successfully",
        data=devices,
        status_code=200
    )


@router.post("/get_device_type", response_model=CustomResponse)
@inject
def get_count(
        device_nature: modelCreate,
        current_user: User = Depends(get_current_active_user),
        device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    models = device_inventory_service.get_device_nature(device_nature)

    return CustomResponse(
        message="Fetched model data successfully",
        data=models,
        status_code=200
    )




@router.post("/get_device_type", response_model=CustomResponse)
@inject
def get_count(
        device_nature: modelCreate,
        current_user: User = Depends(get_current_active_user),
        device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    models = device_inventory_service.get_device_nature(device_nature)

    return CustomResponse(
        message="Fetched model data successfully",
        data=models,
        status_code=200
    )
@router.post("/get_notifications", response_model=CustomResponse)
@inject
def get_expiry(site: site_filter,
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
               ):
    data = device_inventory_service.get_notifications(site)

    return CustomResponse(
        message="Fetched  data successfully",
        data=data,
        status_code=200
    )



@router.post("/generate_excel")
@inject
def generate_excel(filter_data:FilterSchema,
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    devices = device_inventory_service.generate_excel(filter_data)

    print(devices.head(5))
    columns_to_drop = [
        '_sa_instance_state', 'created_at', 'device_id', 'item_desc', 'role', 'contract_number',
        'hardware_version',
        'parent', 'site_id', 'apic_controller_id', 'created_by', 'hw_eol_date',
        'manufacturer_date', 'status', 'sw_eol_date', 'updated_at', 'patch_version',
        'rfs_date', 'sw_eos_date', 'cisco_domain', 'device_ru', 'id', 'criticality', 'hw_eos_date',
        'section', 'tag_id', 'contract_expiry', 'domain', 'modified_by','rack_name','command',
        'source', 'department', 'item_code', 'rack_id', 'manufacturer','software_version','bandwidth_utilization',
        'apic_controller', 'rack', 'site', 'device','pue','performance_score','performance_description'
    ]
    devices = devices.drop(columns=columns_to_drop, errors='ignore')  # errors='ignore' skips missing columns safely
    print(devices.columns,"After drop")

    eer = "Energy Efficiency%\n\nCol Q/Col R\nNote: Power Output/Power Input"
    DEVICE_NAME = "Device Name"
    DEVICE_TYPE = "Device Type"
    SERIAL_NUMBER = "Serial Number"
    PN_CODE = "Product Number (PN)"
    IP_ADDRESS = "IP Address"
    SITE = "Site"
    TOTAL_INTERFACES = "Total Interfaces"
    UP_LINKS = "Up Links"
    DOWN_LINKS = "Down Links"
    ACCESS_PORT = "Access Port"
    UP_LINK_INTERFACES = "UP Link Interfaces"
    INTERFACE_TYPES = "Interface Types"
    STACK = "Stack"
    PSU_COUNT = "PSU Count"
    PSU_MODEL = "PSU Model"
    TOTAL_POWER_CAPACITY = "Total Power Capacity"
    POWER_OUTPUT = "Power Output (W)"
    POWER_INPUT = "Power Input (W)"
    EER_COLUMN = "Energy Efficiency%\n\nCol U(Power Output)/Col V(Power Input)"
    PCR  = "Power Consumption Ratio(W/MB)\n\nCol  V(Power Input) / Col AC(Data Traffic Consumed(MB))\nNote: Power used per input traffic."
    INPUT_PACKETS = "Input Packets"
    OUTPUT_PACKETS = "Output Packets"
    RX_MB = "RX (MB)"
    TX_MB = "TX (MB)"
    DATA_CONSUMED_MB = "Data Traffic Consumed (MB)"
    DATA_CONSUMED_GB = "Data Traffic Consumed (GB)"
    DATA_ALLOCATED_MB = (
        "Data Traffic Allocated (MB)\n\n"
        "Σ {((Interface Bandwidth (in kbps) * 2(duplex)) /(8 * 1024(for kbps to MBps))} *3600(for MBps to MB)\n"
        "Note: Only for Full Duplex we multiply with 2."
    )

    DATA_THROUGHPUT_MB_W = (
        "Data Throughput per Watt (MB/W)\n\n"
        "Col AC((Data Traffic Consumed(MB)) / Col  V(Power Input)\n"
        "Note: Total data traffic handled per watt of power consumed."
    )

    DATA_UTILIZATION_PERCENT = (
        "DataTraffic Utilization%\n\n"
        "Note: Col AC(Data Traffic Consumed)/Col AE(Data Traffic Allocated)"
    )
    CARBON_KG = (
        "Carbon Emission (kgCO₂)\n\n"
        "Note: Power Input(Col V) / 1000(for W to KW) * 0.4041(Carbon Emission Factor)\n"
        "Carbon Emission Factor is Region Specific i.e here UAE"
    )

    CARBON_TON = "Carbon Emission (tonCO₂)"

    HW_EOL = "HW End-of-Life (EoL)"
    HW_EOS = "HW End-of-Support (EoS)"
    HW_LDOS = "HW Last Day of Support"
    HW_RFA = "HW Last Date of RFA"
    HW_SECURITY_EOS = "HW Security Support EoS"
    SW_EOL = "SW Maintenance EoL"
    SW_VULN_EOS = "SW Vulnerability Support EoS"
    ERROR_MSG = "Error Message"
    active_psu ="Active PSU"
    non_active_psu = "Non Active PSU"
    switch_topology = "Switch Topology"
    switch_mode = "Switch Mode"

    new_column_order = [
        # Basic Info
        DEVICE_NAME, DEVICE_TYPE, SERIAL_NUMBER, PN_CODE, IP_ADDRESS, SITE,
        TOTAL_INTERFACES, UP_LINKS, DOWN_LINKS, ACCESS_PORT, UP_LINK_INTERFACES, INTERFACE_TYPES,
        # Power Info
        STACK,active_psu,non_active_psu,switch_topology,switch_mode, PSU_COUNT, PSU_MODEL, TOTAL_POWER_CAPACITY,
        POWER_OUTPUT, POWER_INPUT, EER_COLUMN, PCR,
        # Network
        INPUT_PACKETS, OUTPUT_PACKETS, RX_MB, TX_MB,
        DATA_CONSUMED_MB, DATA_CONSUMED_GB, DATA_ALLOCATED_MB, DATA_THROUGHPUT_MB_W, DATA_UTILIZATION_PERCENT,
        # Environmental
        CARBON_KG, CARBON_TON,
        # Lifecycle
        HW_EOL, HW_EOS, HW_LDOS, HW_RFA, HW_SECURITY_EOS, SW_EOL, SW_VULN_EOS,
        # Errors
        ERROR_MSG
    ]
    devices.rename(columns={
        # Basic Info
        "device_name": DEVICE_NAME,
        "device_type": DEVICE_TYPE,
        "serial_number": SERIAL_NUMBER,
        "pn_code": PN_CODE,
        "device_ip": IP_ADDRESS,
        "site_name": SITE,
        # Interfaces
        "total_interface": TOTAL_INTERFACES,
        "up_link": UP_LINKS,
        "down_link": DOWN_LINKS,
        "access_port": ACCESS_PORT,
        "up_Link_interfaces": UP_LINK_INTERFACES,
        "interfaces_types": INTERFACE_TYPES,
        # Power
        "stack": STACK,
        "active_psu":active_psu,
        "non_active_psu" :non_active_psu,
        "switch_topology" :switch_topology,
        "switch_mode" :switch_mode,
        "psu_count": PSU_COUNT,
        "psu_model": PSU_MODEL,
        "total_power_capacity": TOTAL_POWER_CAPACITY,
        "power_output": POWER_OUTPUT,
        "power_input": POWER_INPUT,
        "power_utilization": EER_COLUMN,
        "pcr": PCR,

        # Network
        "total_input_packets": INPUT_PACKETS,
        "total_output_packets": OUTPUT_PACKETS,
        "total_input_mbs": RX_MB,
        "total_output_mbs": TX_MB,
        "datatraffic": DATA_CONSUMED_MB,
        "datatraffic_gbs": DATA_CONSUMED_GB,
        "bandwidth_mbps": DATA_ALLOCATED_MB,
        "eer_dt": DATA_THROUGHPUT_MB_W,
        "datatraffic_utilization": DATA_UTILIZATION_PERCENT,

        # Environment
        "carbon_emission": CARBON_KG,
        "carbon_emission_tons": CARBON_TON,

        # Lifecycle
        "hw_eol_ad": HW_EOL,
        "hw_eos": HW_EOS,
        "hw_ldos": HW_LDOS,
        "hw_EoRFA": HW_RFA,
        "hw_EoSCR": HW_SECURITY_EOS,
        "sw_EoSWM": SW_EOL,
        "sw_EoVSS": SW_VULN_EOS,
        # Errors
        "error_message": ERROR_MSG
        # "performance_score": "Performance Score",
        # "performance_description": "Performance Description"
    }, inplace=True)
    devices = devices.reindex(columns=new_column_order)

    file_path = "device_report.xlsx"
    # # Save DataFrame to an Excel file
    # devices.to_excel(file_path, index=False, engine="openpyxl")
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        devices.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        max_col = len(new_column_order)

        # Set column widths (customize as needed)
        default_width = 25
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = default_width

        # Wrap text and center align header cells
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Set header row height to fit multiline text
        worksheet.row_dimensions[1].height = 60
    return FileResponse(file_path, filename="device_report.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@router.get("/get_hardware_version",response_model=CustomResponse)
@inject
def get_hardwareversionsa(
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    hardware_versions = device_inventory_service.get_hardware_versions()

    return CustomResponse(
        message="Fetched  hardware versions successfully",
        data=hardware_versions,
        status_code=200
    )
@router.get("/get_software_version",response_model=CustomResponse)
@inject
def get_software_versions(
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    software_versions = device_inventory_service.get_software_versions()

    return CustomResponse(
        message="Fetched  software versions successfully",
        data=software_versions,
        status_code=200
    )


@router.post("/create_vendor/")
@inject
def create_vendor(vendor: VendorSchema, device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    vendor =device_inventory_service.add_vendor(vendor)
    return CustomResponse(
        message="vendor created successfully",
        data=vendor,
        status_code=200
    )


@router.post("/create_device_type/")
@inject
def create_device_type(device: DeviceTypeSchema, device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    device_type =device_inventory_service.add_device_type(device)
    return CustomResponse(
        message="Device Type created successfully",
        data=device_type,
        status_code=200
    )


@router.post("/generate_excel_test")
@inject
def generate_excel_td(filter_data:FilterSchema,
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    devices = device_inventory_service.generate_excel1(filter_data)
    file_path = ".xlsx"
    # Save DataFrame to an Excel file
    devices.to_excel(file_path, index=False, engine="openpyxl")
    return FileResponse(file_path, filename="device_data.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/generate_excel_test1")
@inject
def generate_excel_td(filter_data:FilterSchema,
    # current_user: User = Depends(get_current_active_user),
    device_inventory_service: DeviceInventoryService = Depends(Provide[Container.device_inventory_service])
):
    devices = device_inventory_service.generate_excel1(filter_data)
    file_path = ".xlsx"
    # Save DataFrame to an Excel file
    devices.to_excel(file_path, index=False, engine="openpyxl")
    return FileResponse(file_path, filename="device_data.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

